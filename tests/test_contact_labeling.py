"""
Tests for SEEG contact labeling functionality
"""

import pytest
import pandas as pd
from pathlib import Path
import tempfile
from openpyxl import Workbook

from bidsificator.services.ContactLabelingParser import ContactLabelingParser
from bidsificator.services.BidsMetadataExtractorService import BidsMetadataExtractor


class TestContactLabelingParser:
    """Test ContactLabelingParser functionality"""

    @pytest.fixture
    def sample_excel_file(self):
        """Create a sample Excel file for testing"""
        # Create temporary Excel file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = Path(temp_file.name)
        temp_file.close()

        # Create workbook with sample data
        wb = Workbook()
        ws = wb.active

        # Header row (row 1)
        ws.append([
            'contact',
            'within the EZ', 'Unnamed: 2',
            'within the lesion',
            'spikes (wakefulness)', 'Unnamed: 5',
            'spikes (sleep)', 'Unnamed: 7',
            'ripples (wakefulness)', 'Unnamed: 9',
            'ripples (sleep)', 'Unnamed: 11',
            'fast ripples (wakefulness)', 'Unnamed: 13',
            'fast ripples (sleep)', 'Unnamed: 15',
            'RFTC', 'Resected'
        ])

        # Format hint row (row 2)
        ws.append([
            None, 'y/n/na', 'EI value', None,
            'y/n/nd', 'rate', 'y/n/nd', 'rate',
            'y/n/nd', 'rate', 'y/n/nd', 'rate',
            'y/n/nd', 'rate', 'y/n/nd', 'rate',
            'y/n/na', 'y/n/na'
        ])

        # Data rows
        ws.append(['Y1', 'y', None, None, 'y', 5.2, 'y', 8.1, 'n', None, 'n', None, 'y', 2.3, 'nd', None, 'y', 'y'])
        ws.append(['Y2', 'n', None, None, 'nd', None, 'y', 3.5, 'n', None, 'y', 1.2, 'n', None, 'n', None, 'n', 'n'])
        ws.append(['Y3', 'y', None, None, 'y', 12.0, 'n', None, 'y', 4.5, 'n', None, 'n', None, 'n', None, 'y', 'n'])

        wb.save(temp_path)
        wb.close()

        yield temp_path

        # Cleanup
        temp_path.unlink()

    def test_parse_file_structure(self, sample_excel_file):
        """Test parsing Excel file structure"""
        parser = ContactLabelingParser()
        contact_data = parser.parse_file(sample_excel_file)

        # Should have 3 contacts
        assert len(contact_data) == 3
        assert 'Y1' in contact_data
        assert 'Y2' in contact_data
        assert 'Y3' in contact_data

    def test_parse_file_values(self, sample_excel_file):
        """Test parsing contact annotation values"""
        parser = ContactLabelingParser()
        contact_data = parser.parse_file(sample_excel_file)

        # Check Y1 values
        y1 = contact_data['Y1']
        assert y1['within_ez'] == 'y'
        assert y1['spikes_wake'] == 'y'
        assert y1['spikes_wake_rate'] == 5.2
        assert y1['spikes_sleep'] == 'y'
        assert y1['spikes_sleep_rate'] == 8.1
        assert y1['ripples_wake'] == 'n'
        assert y1['fast_ripples_wake'] == 'y'
        assert y1['rftc'] == 'y'
        assert y1['resected'] == 'y'

        # Check Y2 values
        y2 = contact_data['Y2']
        assert y2['within_ez'] == 'n'
        assert y2['spikes_wake'] == 'nd'
        assert y2['spikes_sleep'] == 'y'
        assert y2['ripples_sleep'] == 'y'
        assert y2['ripples_sleep_rate'] == 1.2

    def test_validate_against_channels(self, sample_excel_file):
        """Test validation against channel names"""
        parser = ContactLabelingParser()
        contact_data = parser.parse_file(sample_excel_file)

        # Test with matching channels
        channel_names = ['Y1', 'Y2', 'Y3', 'Y4']
        validation = parser.validate_against_channels(contact_data, channel_names)

        assert len(validation['matched']) == 3
        assert 'Y1' in validation['matched']
        assert len(validation['missing_in_labeling']) == 1
        assert 'Y4' in validation['missing_in_labeling']
        assert len(validation['missing_in_channels']) == 0

    def test_validate_mismatch(self, sample_excel_file):
        """Test validation with mismatched channel names"""
        parser = ContactLabelingParser()
        contact_data = parser.parse_file(sample_excel_file)

        # Test with different channels
        channel_names = ['A1', 'A2', 'A3']
        validation = parser.validate_against_channels(contact_data, channel_names)

        assert len(validation['matched']) == 0
        assert len(validation['missing_in_channels']) == 3
        assert len(validation['missing_in_labeling']) == 3

    def test_get_annotations_for_contact(self, sample_excel_file):
        """Test getting annotations for specific contact"""
        parser = ContactLabelingParser()
        contact_data = parser.parse_file(sample_excel_file)

        # Test existing contact
        annotations = parser.get_annotations_for_contact(contact_data, 'Y1')
        assert annotations['within_ez'] == 'y'
        assert annotations['spikes_wake_rate'] == 5.2

        # Test non-existing contact
        annotations = parser.get_annotations_for_contact(contact_data, 'Y999')
        assert annotations == {}

    def test_case_insensitive_matching(self, sample_excel_file):
        """Test case-insensitive contact name matching"""
        parser = ContactLabelingParser()
        contact_data = parser.parse_file(sample_excel_file)

        # Test with lowercase channel names (as they appear in EDF files)
        channel_names_lower = ['y1', 'y2', 'y3']
        validation = parser.validate_against_channels(contact_data, channel_names_lower)

        # Should match despite case difference (Y1 vs y1)
        assert len(validation['matched']) == 3
        assert 'y1' in validation['matched']  # Returns in channel case
        assert 'y2' in validation['matched']
        assert 'y3' in validation['matched']
        assert len(validation['missing_in_channels']) == 0
        assert len(validation['missing_in_labeling']) == 0

        # Test case-insensitive annotation lookup
        annotations_lower = parser.get_annotations_for_contact(contact_data, 'y1')
        annotations_upper = parser.get_annotations_for_contact(contact_data, 'Y1')

        # Both should return the same data
        assert annotations_lower == annotations_upper
        assert annotations_lower['within_ez'] == 'y'
        assert annotations_lower['spikes_wake_rate'] == 5.2

        # Test mixed case
        annotations_mixed = parser.get_annotations_for_contact(contact_data, 'y2')
        assert annotations_mixed['within_ez'] == 'n'
        assert annotations_mixed['spikes_sleep_rate'] == 3.5  # Y2 has sleep data, not wake

    def test_invalid_file(self):
        """Test handling of invalid files"""
        parser = ContactLabelingParser()

        # Test non-existent file
        with pytest.raises(FileNotFoundError):
            parser.parse_file(Path('/nonexistent/file.xlsx'))

        # Test non-Excel file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.txt')
        temp_path = Path(temp_file.name)
        temp_file.close()

        try:
            with pytest.raises(ValueError, match="Excel format"):
                parser.parse_file(temp_path)
        finally:
            temp_path.unlink()

    def test_normalize_values(self):
        """Test value normalization"""
        parser = ContactLabelingParser()

        # Test indicator normalization
        assert parser._normalize_value('y') == 'y'
        assert parser._normalize_value('Y') == 'y'
        assert parser._normalize_value('yes') == 'y'
        assert parser._normalize_value('n') == 'n'
        assert parser._normalize_value('N') == 'n'
        assert parser._normalize_value('no') == 'n'
        assert parser._normalize_value('na') == 'n/a'
        assert parser._normalize_value('n/a') == 'n/a'
        assert parser._normalize_value('nd') == 'nd'
        assert parser._normalize_value('n/d') == 'nd'

        # Test numeric values remain unchanged
        assert parser._normalize_value(5.2) == 5.2
        assert parser._normalize_value(10) == 10


class TestBidsMetadataExtractorWithLabeling:
    """Test BidsMetadataExtractor with contact labeling integration"""

    @pytest.fixture
    def sample_electrodes_df(self):
        """Create a sample electrodes DataFrame"""
        data = {
            'name': ['Y1', 'Y2', 'Y3', 'Y4'],
            'x': [10.0, 11.0, 12.0, 13.0],
            'y': [20.0, 21.0, 22.0, 23.0],
            'z': [30.0, 31.0, 32.0, 33.0],
            'size': [1.0, 1.0, 1.0, 1.0],
            'hemisphere': ['L', 'L', 'L', 'L'],
            'group': ['Y', 'Y', 'Y', 'Y']
        }
        return pd.DataFrame(data)

    @pytest.fixture
    def sample_labeling_file(self):
        """Create a sample contact labeling Excel file"""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = Path(temp_file.name)
        temp_file.close()

        wb = Workbook()
        ws = wb.active

        # Simplified header
        ws.append([
            'contact',
            'within the EZ', 'Unnamed: 2',
            'spikes (wakefulness)', 'Unnamed: 5',
            'RFTC', 'Resected'
        ])

        ws.append([None, 'y/n/na', 'EI value', 'y/n/nd', 'rate', 'y/n/na', 'y/n/na'])

        # Data for Y1, Y2, Y3 (no Y4)
        ws.append(['Y1', 'y', None, 'y', 5.2, 'y', 'y'])
        ws.append(['Y2', 'n', None, 'nd', None, 'n', 'n'])
        ws.append(['Y3', 'y', None, 'y', 12.0, 'y', 'n'])

        wb.save(temp_path)
        wb.close()

        yield temp_path

        temp_path.unlink()

    def test_merge_contact_labeling(self, sample_electrodes_df, sample_labeling_file):
        """Test merging contact labeling data into electrodes DataFrame"""
        extractor = BidsMetadataExtractor()

        # Merge labeling data
        merged_df = extractor._merge_contact_labeling_data(
            sample_electrodes_df.copy(),
            sample_labeling_file
        )

        # Check that clinical columns were added
        assert 'within_ez' in merged_df.columns
        assert 'spikes_wake' in merged_df.columns
        assert 'spikes_wake_rate' in merged_df.columns
        assert 'rftc' in merged_df.columns
        assert 'resected' in merged_df.columns

        # Check Y1 values
        y1_row = merged_df[merged_df['name'] == 'Y1'].iloc[0]
        assert y1_row['within_ez'] == 'y'
        assert y1_row['spikes_wake'] == 'y'
        assert y1_row['spikes_wake_rate'] == 5.2
        assert y1_row['rftc'] == 'y'
        assert y1_row['resected'] == 'y'

        # Check Y2 values
        y2_row = merged_df[merged_df['name'] == 'Y2'].iloc[0]
        assert y2_row['within_ez'] == 'n'
        assert y2_row['spikes_wake'] == 'nd'

        # Check Y4 has no labeling data (should be NaN or not present)
        y4_row = merged_df[merged_df['name'] == 'Y4'].iloc[0]
        # Y4 should not have these values or they should be NaN
        assert pd.isna(y4_row.get('within_ez', pd.NA)) or y4_row.get('within_ez') is None

    def test_merge_with_missing_file(self, sample_electrodes_df):
        """Test merging when labeling file doesn't exist"""
        extractor = BidsMetadataExtractor()

        # Should return original DataFrame without error
        merged_df = extractor._merge_contact_labeling_data(
            sample_electrodes_df.copy(),
            Path('/nonexistent/file.xlsx')
        )

        # Should be same as input (no clinical columns)
        assert len(merged_df.columns) == len(sample_electrodes_df.columns)


class TestEndToEndIntegration:
    """Test end-to-end integration through BidsSubject workflow"""

    @pytest.fixture
    def sample_labeling_file(self):
        """Create a sample contact labeling Excel file for end-to-end test"""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_path = Path(temp_file.name)
        temp_file.close()

        wb = Workbook()
        ws = wb.active

        # Header
        ws.append([
            'contact',
            'within the EZ', 'Unnamed: 2',
            'spikes (wakefulness)', 'Unnamed: 5',
            'RFTC', 'Resected'
        ])

        ws.append([None, 'y/n/na', 'EI value', 'y/n/nd', 'rate', 'y/n/na', 'y/n/na'])

        # Sample data
        ws.append(['CH001', 'y', None, 'y', 5.2, 'y', 'y'])
        ws.append(['CH002', 'n', None, 'nd', None, 'n', 'n'])

        wb.save(temp_path)
        wb.close()

        yield temp_path

        temp_path.unlink()

    def test_extract_electrodes_with_labeling(self, sample_labeling_file):
        """Test that extract_electrodes_tsv properly merges labeling data"""
        extractor = BidsMetadataExtractor()

        # Create mock electrodes data (simulating what would be extracted from file)
        # In real scenario, this would come from converter
        mock_electrodes_raw = [
            {'name': 'CH001', 'x': 10.0, 'y': 20.0, 'z': 30.0},
            {'name': 'CH002', 'x': 11.0, 'y': 21.0, 'z': 31.0}
        ]

        # Create base electrodes DataFrame (simulates converter output)
        base_df = pd.DataFrame(mock_electrodes_raw)

        # Now merge with labeling file
        merged_df = extractor._merge_contact_labeling_data(
            base_df.copy(),
            sample_labeling_file
        )

        # Verify clinical columns are present
        assert 'within_ez' in merged_df.columns
        assert 'spikes_wake' in merged_df.columns
        assert 'spikes_wake_rate' in merged_df.columns

        # Verify CH001 has correct values
        ch001_row = merged_df[merged_df['name'] == 'CH001'].iloc[0]
        assert ch001_row['within_ez'] == 'y'
        assert ch001_row['spikes_wake'] == 'y'
        assert ch001_row['spikes_wake_rate'] == 5.2
        assert ch001_row['rftc'] == 'y'
        assert ch001_row['resected'] == 'y'

        # Verify CH002 has correct values
        ch002_row = merged_df[merged_df['name'] == 'CH002'].iloc[0]
        assert ch002_row['within_ez'] == 'n'
        assert ch002_row['spikes_wake'] == 'nd'

    def test_contact_labeling_file_validation(self, sample_labeling_file):
        """Test that BidsSubject properly validates and stores labeling file"""
        from bidsificator.core.BidsSubjectSchema import BidsSubject
        from bidsificator.core.schema import BidsSchemaManager
        import tempfile

        # Create temporary dataset directory
        with tempfile.TemporaryDirectory() as tmpdir:
            dataset_path = Path(tmpdir)
            schema_manager = BidsSchemaManager.get_instance()

            # Create subject
            subject = BidsSubject('test01', dataset_path, schema_manager)

            # Initially no labeling file
            assert subject.get_contact_labeling_file() is None
            assert not subject.has_contact_labeling_file()

            # Set labeling file
            subject.set_contact_labeling_file(sample_labeling_file)

            # Now should have labeling file
            assert subject.get_contact_labeling_file() == sample_labeling_file
            assert subject.has_contact_labeling_file()

            # Test validation - non-existent file should raise error
            with pytest.raises(FileNotFoundError):
                subject.set_contact_labeling_file(Path('/nonexistent/file.xlsx'))

            # Test validation - non-Excel file should raise error
            txt_file = Path(tmpdir) / 'test.txt'
            txt_file.write_text('test')
            with pytest.raises(ValueError, match="Excel format"):
                subject.set_contact_labeling_file(txt_file)
